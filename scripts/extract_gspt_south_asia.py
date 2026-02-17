"""
Extract and filter GSPT (Global Solar Power Tracker) data for South Asia.

Reads the GSPT Excel file, filters for South Asian countries,
and outputs a JSON file with key project fields.
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

DATA_DIR = Path(__file__).parent.parent / "data"
GSPT_FILE = DATA_DIR / "Global-Solar-Power-Tracker-February-2026.xlsx"
OUTPUT_FILE = DATA_DIR / "gspt_south_asia.json"

SOUTH_ASIA_COUNTRIES = {"India", "Bangladesh", "Pakistan", "Bhutan", "Nepal", "Sri Lanka"}

# Column mapping (1-indexed from Excel)
COLUMNS = {
    "date_last_researched": 1,
    "country": 2,
    "project_name": 3,
    "phase_name": 4,
    "capacity_mw": 7,
    "capacity_rating": 8,
    "technology_type": 9,
    "status": 10,
    "start_year": 11,
    "retired_year": 12,
    "operator": 13,
    "owner": 15,
    "latitude": 19,
    "longitude": 20,
    "location_accuracy": 21,
    "state_province": 25,
    "gem_location_id": 28,
    "gem_phase_id": 29,
    "other_ids": 30,
    "wiki_url": 32,
}


def extract_gspt():
    print(f"Reading {GSPT_FILE}...")
    wb = openpyxl.load_workbook(str(GSPT_FILE), read_only=True, data_only=True)
    ws = wb["Utility-Scale (1 MW+)"]

    # Read header row to verify column mapping
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"Found {len(header)} columns")
    print(f"Sample headers: {header[:5]}...")

    projects = []
    skipped = 0
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        country = row[COLUMNS["country"] - 1]
        if country not in SOUTH_ASIA_COUNTRIES:
            skipped += 1
            continue

        # Extract fields
        project = {}
        for field, col_idx in COLUMNS.items():
            val = row[col_idx - 1]
            # Convert numeric types
            if field in ("capacity_mw", "latitude", "longitude"):
                val = float(val) if val is not None else None
            elif field in ("start_year", "retired_year"):
                val = int(val) if val is not None else None
            else:
                val = str(val).strip() if val is not None else None
            project[field] = val

        projects.append(project)

    wb.close()

    print(f"\nProcessed {total} rows total")
    print(f"Filtered to {len(projects)} South Asia projects")
    print(f"Skipped {skipped} non-South-Asia rows")

    # Summary statistics
    countries = {}
    statuses = {}
    with_coords = 0
    for p in projects:
        countries[p["country"]] = countries.get(p["country"], 0) + 1
        statuses[p["status"]] = statuses.get(p["status"], 0) + 1
        if p["latitude"] is not None and p["longitude"] is not None:
            with_coords += 1

    print(f"\nBy country:")
    for c, n in sorted(countries.items(), key=lambda x: -x[1]):
        print(f"  {c}: {n}")

    print(f"\nBy status:")
    for s, n in sorted(statuses.items(), key=lambda x: -x[1]):
        print(f"  {s}: {n}")

    print(f"\nWith coordinates: {with_coords} / {len(projects)}")

    # Save
    with open(OUTPUT_FILE, "w") as f:
        json.dump(projects, f, indent=2, default=str)
    print(f"\nSaved to {OUTPUT_FILE}")
    print(f"File size: {OUTPUT_FILE.stat().st_size / 1024:.0f} KB")

    return projects


if __name__ == "__main__":
    extract_gspt()
